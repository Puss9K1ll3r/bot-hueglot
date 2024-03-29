import os
import openpyxl
from zipfile import ZipFile
#
def GetFileName():
  excelFileName = ""
  content = os.listdir()
  for i in range(len(content)):
    if ".xlsx" in content[i]:
      excelFileName = content[i]
    elif ".xls" in content[i]:
      excelFileName = content[i]
  return excelFileName


def GetIndexesOfDays(sheet):
  daysOfWeek = [
    "понедельник", "вторник", "среда", "четверг", "пятница", "суббота"
  ]
  daysNum = 0
  dayIndex = list()
  for row in range(1, sheet.max_row + 1):
    if sheet[row][0].value:
      if daysOfWeek[daysNum] in sheet[row][0].value.lower():
        daysNum += 1
        dayIndex.append(row)
      if daysNum > 5:
        break
  row = dayIndex[-1]
  while sheet[row][1].value:
    row += 1
  dayIndex.append(row)
  return dayIndex


def Redactor(scheduleContent):
  dCNameValue = True
  text = list()
  for dCNum, dCName, audithory in scheduleContent:
    if dCName.value:
      dCNameValue = dCName.value
      if "\n" in dCNameValue:
        dCNameValue = dCNameValue.replace("\n", "\n\n")
        
      if audithory.value:
        audithoryValue = audithory.value
        if "с/з" in audithoryValue:
          audithoryValue = audithoryValue.replace("с/з,", "с/з")
        audithory = audithoryValue.replace("\n", "").split(",")[:-1]
        audithory = "\n".join(audithory)
        
        text.append([dCNum.value, dCNameValue, "___________", audithory])
      else:
        text.append([dCNum.value, dCNameValue, "___________", ""])
    else:
      text.append([dCNum.value, "___________", ""])
  return text


def ListJoin(text, fileName):
  for i in range(len(text)):
    text[i] = "\n".join(map(str, text[i]))
  text = "\n\n".join(map(str, text))
  with open(fileName, "w") as openedFile:
    openedFile.write(text)


def ZipFileOpen(excelName, filesWithSchedule):
  if ".xlsx" in excelName:
    archiveName = excelName[:-5] + ".zip"
  else:
    archiveName = excelName[:-4] + ".zip"
  with ZipFile(archiveName, "w") as archive:
    for i in range(len(filesWithSchedule)):
      archive.write(filesWithSchedule[i])


def RemoveFiles(scheduleFiles, scheduleXlsxFile):
  for i in range(len(scheduleFiles)):
    os.remove(scheduleFiles[i])
  os.remove(scheduleXlsxFile)


def RemoveOldScheduleArchives(archiveName):
  if ".xlsx" in archiveName:
    archiveName = archiveName[:-5] + ".zip"
  else:
    archiveName = archiveName[:-4] + ".zip"
  content = os.listdir()
  for i in range(len(content)):
    if ".zip" in content[i]:
      if archiveName[:3] in content[i] and archiveName != content[i]:
        os.remove(content[i])
      else:
        continue


def ScheduleCreate():
  try:
    scheduleFileName = GetFileName()
    file = openpyxl.open(scheduleFileName, read_only=True)
  except:
    print("Excel files didnt found in the directory, please check it.")
  else:
    sheet = file.active
    indexesOfDays = GetIndexesOfDays(sheet)
    daysOfWeek = [
      "понедельник", "вторник", "среда", "четверг", "пятница", "суббота"
    ]
    scheduleFiles = list()

    for i in range(len(indexesOfDays) - 1):
        if daysOfWeek[0] in sheet[indexesOfDays[i]][0].value.lower():
          schedule = sheet.iter_rows(min_row=indexesOfDays[i], max_row=(indexesOfDays[i + 1] - 1), min_col=2, max_col=4)
          ListJoin(Redactor(schedule), "monday.txt")
          scheduleFiles.append("monday.txt")
        elif daysOfWeek[1] in sheet[indexesOfDays[i]][0].value.lower():
          schedule = sheet.iter_rows(min_row=indexesOfDays[i], max_row=(indexesOfDays[i + 1] - 1), min_col=2, max_col=4)
          ListJoin(Redactor(schedule), "tuesday.txt")
          scheduleFiles.append("tuesday.txt")
        elif daysOfWeek[2] in sheet[indexesOfDays[i]][0].value.lower():
          schedule = sheet.iter_rows(min_row=indexesOfDays[i], max_row=(indexesOfDays[i + 1] - 1), min_col=2, max_col=4)
          ListJoin(Redactor(schedule), "wednesday.txt")
          scheduleFiles.append("wednesday.txt")
        elif daysOfWeek[3] in sheet[indexesOfDays[i]][0].value.lower():
          schedule = sheet.iter_rows(min_row=indexesOfDays[i], max_row=(indexesOfDays[i + 1] - 1), min_col=2, max_col=4)
          ListJoin(Redactor(schedule), "thursday.txt")
          scheduleFiles.append("thursday.txt")
        elif daysOfWeek[4] in sheet[indexesOfDays[i]][0].value.lower():
          schedule = sheet.iter_rows(min_row=indexesOfDays[i], max_row=(indexesOfDays[i + 1] - 1), min_col=2, max_col=4)
          ListJoin(Redactor(schedule), "friday.txt")
          scheduleFiles.append("friday.txt")
        elif daysOfWeek[5] in sheet[indexesOfDays[i]][0].value.lower():
          schedule = sheet.iter_rows(min_row=indexesOfDays[i], max_row=(indexesOfDays[i + 1] - 1), min_col=2, max_col=4)
          ListJoin(Redactor(schedule), "saturday.txt")
          scheduleFiles.append("saturday.txt")
    file.close()

    RemoveOldScheduleArchives(GetFileName())

    ZipFileOpen(GetFileName(), scheduleFiles)

    RemoveFiles(scheduleFiles, scheduleFileName)